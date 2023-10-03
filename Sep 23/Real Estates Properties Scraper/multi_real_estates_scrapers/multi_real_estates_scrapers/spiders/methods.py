import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook  # Load  worksheet
from openpyxl.utils import get_column_letter  # Font Color
from openpyxl.workbook import Workbook  # Worksheet related Tasks


def comparison_excel(data):
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)
    file_name = f'{output_dir} Output Data.xlsx'

    try:
        if os.path.isfile(file_name):
            wb = load_workbook(file_name)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)
        scrapedata_headers = ['Address', 'street number', 'Type', 'Rooms', 'Other', 'Size (m2)', 'Agency url',
                              'Agency name']

        fulldata_sheet_name = 'FULLDATA'
        if fulldata_sheet_name in wb.sheetnames:
            fulldata_sheet = wb[fulldata_sheet_name]
        else:
            fulldata_sheet = wb.create_sheet(fulldata_sheet_name)
            fulldata_headers = ['Address', 'street number', 'Other', 'Size (m2)', 'Agency url', 'Agency name']
            fulldata_sheet.append(fulldata_headers)

        scrapedata_sheet_name = 'SCRAPEDATA'
        if scrapedata_sheet_name in wb.sheetnames:
            scrapedata_sheet = wb[scrapedata_sheet_name]
        else:
            scrapedata_sheet = wb.create_sheet(scrapedata_sheet_name)
            scrapedata_sheet.append(scrapedata_headers)

        addresses_sheet_name = 'ADDRESSES'
        if addresses_sheet_name in wb.sheetnames:
            addresses_sheet = wb[addresses_sheet_name]
        else:
            addresses_sheet = wb.create_sheet(addresses_sheet_name)
            addresses_headers = ['Address', 'street number']
            addresses_sheet.append(addresses_headers)

        storage_sheet_name = 'STORAGE'
        if storage_sheet_name in wb.sheetnames:
            storage_sheet = wb[storage_sheet_name]
        else:
            storage_sheet = wb.create_sheet(storage_sheet_name)
            # storage_headers = ['Address', 'street number']
            storage_headers = ['Address', 'street number', 'Other', 'Size (m2)', 'Agency url', 'Agency name', 'timestamp']
            storage_sheet.append(storage_headers)

        timestamp_format = "%Y-%m-%d %H:%M:%S"

        # *********step 1  ***************

        fulldata_headers = [cell.value for cell in fulldata_sheet[1]]
        if fulldata_headers and 'Agency url' in fulldata_headers:

            # Remove all rows in FULLDATA worksheet where there is any content in column S (agency_url)
            column_agency_url = fulldata_headers.index("Agency url") + 1  # Adding 1 because column indexes are 1-based
            fulldata_rows_to_delete = []

            for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row, min_col=column_agency_url,
                                                max_col=column_agency_url):
                if row[0].value:
                    fulldata_rows_to_delete.append(row[0].row)

            for row_number in reversed(fulldata_rows_to_delete):
                fulldata_sheet.delete_rows(idx=row_number)

        # ********* Step 2: Remove All Rows in SCRAPEDATA ***************
        if 'SCRAPEDATA' in wb.sheetnames:
            scrapedata_sheet = wb['SCRAPEDATA']
            scrapedata_sheet.delete_rows(2, scrapedata_sheet.max_row)
            print('All Rows Deleted from SCRAPEDATA Sheet Successfully')

        # *********  Add Scrape Results from Spiders into SCRAPEDATA worksheet ***************
        for item in data:
            if not item.get('Agency url', ''):
                continue  # Skip this item if 'Agency url' is empty
            scrapedata_sheet.append([item.get(col) for col in scrapedata_headers])

        # ********* Step 3  ***************

        # Check if the headers are in fulldata_headers, and if not, use scrapedata_headers
        if "Address" in fulldata_headers and "street number" in fulldata_headers:
            column_address = fulldata_headers.index("Address") + 1  # Adding 1 because column indexes are 1-based
            column_street_number = fulldata_headers.index("street number") + 1
        else:
            column_address = scrapedata_headers.index("Address") + 1  # Use scrapedata_headers
            column_street_number = scrapedata_headers.index("street number") + 1

        # Iterates row by row from Scrapedata worksheet for next processing.
        for scrapeddata_row in scrapedata_sheet.iter_rows(min_row=2, max_row=scrapedata_sheet.max_row):
            scrapeddata_values = [cell.value for cell in scrapeddata_row]

            # Start Checking if columns Address(B) and Street No (C) in SCRAPEDATA match any row in FULLDATA
            matching_row = None

            # check for fulldata sheet is not empty
            if fulldata_headers and column_address <= len(fulldata_headers) and column_street_number <= len(fulldata_headers):
                for fulldata_row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                                             min_col=column_address, max_col=column_street_number):
                    fulldata_values = [cell.value for cell in fulldata_row]

                    # If scrapedata rows matched in fulldata sheet address and street no
                    if fulldata_values == scrapeddata_values[0:2]:  # comparison address and street no, 0 to 2
                        matching_row = fulldata_row[0].row

                        # Loop through scrapeddata headers and values to insert into corresponding columns in FULLDATA
                        for header, value in zip(scrapedata_headers, scrapeddata_values):
                            if header in fulldata_headers:
                                col_index = fulldata_headers.index(header) + 1  # Find the column index for the header
                                fulldata_sheet.cell(row=matching_row, column=col_index, value=value)
                        break

            # Check if it's still None (no match in FULLDATA)
            if matching_row is None:
                print(f'Scrapedata row {scrapeddata_values} not matched in FULLDATA worksheet, now check in ADDRESSES worksheet')

                # Initialize a flag to track whether the row is matched in the addresses worksheet
                matched_in_addresses = False

                for addresses_row in addresses_sheet.iter_rows(min_row=2, max_row=addresses_sheet.max_row,
                                                               min_col=column_address, max_col=column_street_number):
                    addresses_values = [cell.value for cell in addresses_row]
                    if addresses_values == scrapeddata_values[0:2]:

                        data_to_append = []

                        for header in fulldata_headers:
                            if header in scrapedata_headers:
                                col_index = scrapedata_headers.index(header) + 1
                                if col_index <= len(scrapeddata_values):
                                    data_to_append.append(scrapeddata_values[col_index - 1])
                                else:
                                    data_to_append.append(
                                        None)  # Append None if no matching value found in scrapeddata_values
                            else:
                                data_to_append.append(None)  # Append None for headers that don't match

                        # Append the data to FULLDATA
                        fulldata_sheet.append(data_to_append)
                        matched_in_addresses = True
                        break

                if not matched_in_addresses:
                    print(f'Scrapedata row {scrapeddata_values} not matched in Addresses worksheet and Ignored.')

        # Copy data from FULLDATA to STORAGE with duplicates checked.
        existing_data = {}
        for storage_row in storage_sheet.iter_rows(min_row=2, max_row=storage_sheet.max_row,
                                                   min_col=column_address, max_col=column_street_number):
            key = tuple(cell.value for cell in storage_row[column_address - 1:column_street_number])
            existing_data[key] = [cell.value for cell in storage_row]

        # Iterate through rows in FULLDATA
        for row in fulldata_sheet.iter_rows(min_row=2, max_row=fulldata_sheet.max_row,
                                            min_col=1, max_col=len(fulldata_headers)):

            fulldata_values = [cell.value for cell in row]
            address_street_key = tuple(fulldata_values[column_address - 1:column_street_number])

            # Check if the row already exists in STORAGE and if it differs from the new data
            if address_street_key in existing_data and existing_data[address_street_key] != fulldata_values:
                timestamp = datetime.now().strftime(timestamp_format)
                storage_sheet.append(fulldata_values + [timestamp])
            elif address_street_key not in existing_data:
                timestamp = datetime.now().strftime(timestamp_format)
                storage_sheet.append(fulldata_values + [timestamp])
                existing_data[address_street_key] = fulldata_values

        # Set column widths for All sheet
        column_widths = {
            'Address': 25,
            'street number': 15,
            'Type': 20,
            'Rooms': 10,
            'Other': 10,
            'Size (m2)': 15,
            'Agency url': 50,
            'Agency name': 15,
            'timestamp': 25,
        }

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for header_name, width in column_widths.items():
                for column_letter, cell in enumerate(sheet[1], start=1):
                    if cell.value == header_name:
                        if width is not None:
                            sheet.column_dimensions[get_column_letter(column_letter)].width = width

        wb.save(file_name)
        print(f" Successfully create output File :{file_name}")

    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")

    return file_name
