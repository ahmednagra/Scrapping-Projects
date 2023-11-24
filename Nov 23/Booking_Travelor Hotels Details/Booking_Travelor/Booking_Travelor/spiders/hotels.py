import csv
import json
import re
import time
from math import ceil
from time import sleep
from datetime import datetime

import requests
from collections import OrderedDict
from scrapy import Spider, Request

from scrapy import Selector
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.options import Options as EdgeOptions

from msedge.selenium_tools import EdgeOptions

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from itertools import zip_longest


class HotelsSpider(Spider):
    name = 'hotels'
    # start_urls = ['https://www.booking.com']

    custom_settings = {
        'CONCURRENT_REQUESTS': 3,
        # 'FEEDS': {
        #     f'output/Hotels Booking Details.csv': {
        #         'format': 'csv',
        #         'fields': ['Hotel Name', 'Location', 'Price', 'Rooms', 'Date Checkin', 'Date Checkout', 'Adults',
        #                    'Children', 'URL'],
        #         'overwrite': True
        #     }
        # },
    }

    booking_cityid_headers = {
        'authority': 'accommodations.booking.com',
        'accept': '*/*',
        'accept-language': 'en-PK,en;q=0.9',
        'content-type': 'text/plain;charset=UTF-8',
        # 'cookie': 'px_init=0; pcm_consent=analytical%3Dtrue%26countryCode%3DPK%26consentId%3Dceb63fcc-9fc2-464a-89e8-229701ae6b6d%26consentedAt%3D2023-11-13T08%3A16%3A55.902Z%26expiresAt%3D2024-05-11T08%3A16%3A55.902Z%26implicit%3Dtrue%26marketing%3Dtrue%26regionCode%3DPB%26regulation%3Dnone%26legacyRegulation%3Dnone; bkng_sso_auth=CAIQsOnuTRpyM330/TpJy9/swUjKG0V6t4ZQfUeRVtEaQ7x2ACm05HLVBKSFaWEOE7XCEDteTA1MqY56YXrLiEN0CyoKpVoVP27yzfbh00EO51IAnLKxdq1MDmXMgntWSRfyWJ8Wh0W1HsfpRa37w+Qjn2bPAipQ3a2P; _pxhd=WLV7jAuFnLvRd7KevLNFUqOfAAybv8HhswRPDQK9AYdzJNI1462CNnPsXhf0Qfwo8WF2xr68XEcvNUhvLbfhhA%253D%253D%253A5ZPT115e6xp7Ht-gTq6wFtqDTvU%252FE5HNincrSIZRUex0n07XJ7AXKlyh%252FnDYclhCgZh3UMDq06FryB5TwYGrjjtrpbSe4w3nr9v7XoV-cSU%253D; cors_js=1; BJS=-; _gid=GA1.2.1144628141.1699863422; bkng_sso_session=e30; bkng_sso_ses=e30; pxcts=07bf9a36-81fd-11ee-ac8a-cf50c8eb43aa; _pxvid=07bf8a32-81fd-11ee-ac8a-3c37de9a2aca; bkng_prue=1; _gcl_au=1.1.1036311762.1699863428; FPID=FPID2.2.Qt5%2Bqi6eh8OM5oLoqx492bmqP%2BP9NVAPExcjNmbeOQA%3D.1699863422; FPLC=RSHnydQvqd2HYz7U8ILsmgaMURGZFi2P6891ZkmTdTSzMxJGHn%2ByYEf9EGiz%2BkqIG0LYC98efGByZ0uE4yn5z2UqbZxXClEbqxsNkOS%2FIIXy8UDx%2FzK0eW7yDqCQsQ%3D%3D; _scid=dc849ed5-3007-4956-b6e2-336249e3414a; _sctr=1%7C1699815600000; _pin_unauth=dWlkPVl6QTRaamc0WlRBdE56QTJNeTAwWXpRMExXRTJZVFV0WVROaU16TTNOMkZoTWpnMg; _gat=1; _ga_FPD6YLJCJ7=GS1.1.1699863433.1.1.1699864629.60.0.0; _ga=GA1.1.378366241.1699863422; _ga_A12345=GS1.1.1699863433.1.1.1699864629.0.0.0; _scid_r=dc849ed5-3007-4956-b6e2-336249e3414a; _px3=6a5a73afc2f8ab3dfa8bc2fd0d31deceb853ce7ef894dc0d7b35c2b53168bdce:+fODFmb31KKl6v5T+5aTZw3/VuMyP5D7d0p7t/kQgMXlVxR7P1+3JH35hNyB+wHwQgjMOQOPUGKFWRZ0KRA/Ew==:1000:Z0sHEQ2qMb34OcUvFh5/jlAA2/ub+r9igkuiwoGaKKD012f+m4HHeM2pTvSrjTA9w/NICe1GSWZLYns+XGDJELmfwC722sWbIZsolXC+diiQv8gmWsvB/Mn0TPqk3c/r5gCRtOohJgjReT1xKO5QpKEXhtRJRNvPlZ6gQCX3q3AjlrEN+YbxbDL5nsN3psU9wWNcdYJgN4/kbpv63Ajk/v5295IaZ9Bpdg/dlNNPROc=; _pxde=2273f8d12a5cfe2fce829c8fdc73f4a3cfdaa4a2b752c1ba34bb2e164c237d72:eyJ0aW1lc3RhbXAiOjE2OTk4NjQ2MzAwMTcsImZfa2IiOjAsImlwY19pZCI6W119; _uetsid=0e3e1ee081fd11ee8f72d775f1172005; _uetvid=0e3ef50081fd11eeb797d9641f0f24a1; bkng=11UmFuZG9tSVYkc2RlIyh9Yaa29%2F3xUOLbca8KLfxLPeerMxWo2fXXM3nyDj22mbohQmqK50Hw6ynij%2F5kx7Ec7pcTz7EFuv19dUMu1gTYbYa02dOe7su5LRwtD8EcIiP77vSmV7DawYXo6aduv8iEe5Qt03C7zNPbnzEya4HNDNpeko8hV8l8K%2B910YEaKbStxoJ2bR7%2FinI%3D; lastSeen=0',
        'origin': 'https://www.booking.com',
        'referer': 'https://www.booking.com/',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    }

    travelor_placeid_headers = {
        'authority': 'maps.googleapis.com',
        'accept': '*/*',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        'referer': 'https://www.travelor.com/',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'script',
        'sec-fetch-mode': 'no-cors',
        'sec-fetch-site': 'cross-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
        'x-client-data': 'CIq2yQEIorbJAQipncoBCM7kygEIlaHLAQiLq8wBCIWgzQEI6MXNAQi71M0BGOmYzQE=',
    }

    booking_headers = {
        'authority': 'www.booking.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'en-PK,en;q=0.9,ur-PK;q=0.8,ur;q=0.7,en-US;q=0.6',
        # 'cookie': 'px_init=0; pcm_consent=analytical%3Dtrue%26countryCode%3DPK%26consentId%3Dd441d507-4146-47e3-93a7-905b1f3d6d8f%26consentedAt%3D2023-11-13T07%3A41%3A48.421Z%26expiresAt%3D2024-05-11T07%3A41%3A48.421Z%26implicit%3Dtrue%26marketing%3Dtrue%26regionCode%3DPB%26regulation%3Dnone%26legacyRegulation%3Dnone; _pxhd=wyXJL0LF6kuoj-m%252Fy1SNqDjfbTymhUyycNY1bQuaK9blEmX5PwN5PBbcZQQouZ6jaJnjUSscHWHO3CntEDu%252F%252FA%253D%253D%253APL1AF%252F3peWhQyNJ6lwtOtCIgVKLJuptVjR4K6xbPtaoZd2SB2ofpBxwRVv--C1ler%252FlaQY3hI1Og91MJdtPAotKv1lXIiH8K90q0LiHQP6Q%253D; cors_js=1; bkng_sso_session=e30; bkng_sso_ses=e30; OptanonConsent=implicitConsentCountry=nonGDPR&implicitConsentDate=1699861319217; BJS=-; _gid=GA1.2.1899408764.1699861332; pxcts=28a86fa6-81f8-11ee-9064-479dd33721c1; _pxvid=28a8638e-81f8-11ee-9064-161d059c0eeb; bkng_prue=1; _gcl_au=1.1.506763756.1699861340; _scid=fdbb7560-dc0b-4009-b734-a8cdf26ca522; FPID=FPID2.2.1HyBvDbVuYxl01GquN4uBfDD4CpcMluRfS5u3nZi7bI%3D.1699861332; _sctr=1%7C1699815600000; _pin_unauth=dWlkPU9UbGpaR1E0WkRjdE9EazJaQzAwWVRrekxUa3daalF0TmpFNVpHTm1ZbU0xWTJObA; bkng_sso_auth=CAIQsOnuTRpy8ZEQua6ZoaTtTUVCY3DX2JAsv8jWKcPZ6W00bUCBwRAkQiYzvfCfeKQq10j0p5i4+mbrKtM0NQk1CerT4gN/UorkbZRb4GcaubvL/f42bN/Gd826NqqCzeAy9+C+ljrIaK2ZKrD7ts3gPEs4G4aj9KiL; g_state={"i_p":1700020529427,"i_l":2}; FPLC=KNkCiziH%2FSQaG7hReLD57Ey3oCCIVQ4A1ye1y%2F7iCEeqz%2FYBvAC2ZzmAp5ujp1Iosc%2Be1MSgVzdqpf0CodN7y89paZQez3%2Fay0U6WfBNZtFWWXVQnJ1PD0jyQpB%2FDg%3D%3D; cto_bundle=y28tJ19IMXhLNkk1dnJuUWpkSkZ3JTJGUm94VW5mSHNRUlNsUzZxcmdFR2FTUW5iR2pEZE1DUDNYZnl3WlUlMkJ5U1Badjc1aEs4WVI5NnVvTFpXZ0FlSXhCRUdWMWl5VzRObDdkblpZbENBJTJGRHhEeSUyRmM3M0dSYTlXb2NzMDFEM05aOCUyRlgydE13aXVHQ0F1NXpwNDNjbDFGcEVTSERrZVVaY1d0ZnFBMGlCJTJGc0FsR2NxenVKeiUyRiUyQmlTYk5yamtDeVo1N0pqTTlBSVphUFQlMkJMcWtXWDhoc1olMkJmUUpXQVElM0QlM0Q; _pxff_cfp=1; _pxff_ddtc=1; _pxff_tm=1; _gat=1; _px3=b675999480a7f0d8bb896a26a3e3fdf183f419586851aa430ffeef02d6422eaf:Gh5TC1bblj+g4A3ZHjeu3aF2Xw438Q3uohGCDjrKbDKYaYZXJ4rzHUCx7qTcnrut2/s48X9W/DPs26VTj4CW8w==:1000:3u3VqvLUJ1ZduBmR+80A1AlBerluRsBfAh0agQat6mlGgIzrvCdAKaWTsgM8bOszg0uNnn13sKgEEKVGGZRhEiVX/orK397sk1LJMTi/i/fVln3yBTJyqzbj18hJ0kMJToNRlNJNEqxsw8Q5CNOhqXyCFVrbVUicBI03ENn3zSGfPoRcr83QTXUpLvr3za/gKwnQjcOE7SpFTR0qjWkgWNTAMj/SsFwzmi3ouUfo2/8=; _pxde=a0b7c6533e41ea584597b620e149c1b944e68d30675771d9e22dc40878a3a975:eyJ0aW1lc3RhbXAiOjE2OTk5MzQyMzI5MTMsImZfa2IiOjAsImlwY19pZCI6W119; bkng=11UmFuZG9tSVYkc2RlIyh9YfDNyGw8J7nzPnUG3Pr%2Bfv63Tuq2%2FgpOvVyl0QGsHptE4YzG4e6LtBaFZz1QujnlJ1MHrPxMeRk%2FLIWuTgprssrqPzaKH9bzTgLsgkaHh6pZt7WXSKkhPNHyTsDL2QLihK6zX98koTkO0AoP54uwDBRZnqviL057rz5ltB2SCdNf2K22X3BnPoJzOY%2BADiOsMg%3D%3D; _ga_A12345=GS1.1.1699934143.2.1.1699934235.0.0.0; _ga=GA1.1.1260210160.1699861332; _uetsid=33f57ff081f811eea0dffbfae6f907ee; _uetvid=33f5fc6081f811ee9eec5fcab8e29b33; _screload=1; _scid_r=fdbb7560-dc0b-4009-b734-a8cdf26ca522; lastSeen=0; _ga_FPD6YLJCJ7=GS1.1.1699934143.2.1.1699934245.36.0.0',
        # 'referer': 'https://www.booking.com/index.en-gb.html?label=gen173nr-1FCAQoggI49ANICVgEaLUBiAEBmAEJuAEXyAEM2AEB6AEB-AECiAIBqAIDuAKS38uqBsACAdICJDRjYmEwZDRmLTdjODYtNDQ0Yy1iOTZkLWE3ZmMyZWU1ODJjONgCBeACAQ&aid=304142&sid=170d1627059be9343d53beba5e3283d6',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    }

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.bookinginfo = self.get_key_values_from_file('input/bookinginfo.txt')
        self.booking_scraped_items = []
        self.travelor_scraped_items = []

        self.scraped_urls = []
        self.driver = None  # Initialize the driver as None

    def start_requests(self):
        urls = ['https://www.booking.com', 'https://www.travelor.com/he?fid=68193']
        # urls = ['https://www.booking.com']
        for url in urls:
            if 'booking' in url:
                yield Request(url=url, callback=self.parse)

            else:
                yield Request(url=url, callback=self.travelor_parse)

    def parse(self, response, **kwargs):
        city_id = self.booking_get_cityid(response)
        booking_url = self.booking_query_url(response, city_id)

        yield Request(url=booking_url, callback=self.booking_pagination, headers=self.booking_headers)

    def booking_pagination(self, response):
        total_hotels = response.css('h1[aria-live="assertive"] ::text').get('')
        hotels = int(''.join(re.findall(r'\d+', total_hotels)))
        total_pages = ceil(hotels / 25)

        for page in range(1, total_pages + 1):
            offset = 25 * (page - 1)
            url = response.url + f'&offset={offset}'
            yield Request(url=url, headers=self.booking_headers,
                          callback=self.booking_parse_hotels)

    def booking_parse_hotels(self, response):

        hotels_selectors = response.css('[data-testid="property-card"]')
        for hotel in hotels_selectors:
            item = OrderedDict()
            string = hotel.css('[data-testid="price-for-x-nights"] ::text').getall() or []
            url = hotel.css('[data-testid="title-link"]::attr(href)').get('')

            item['Search Location'] = self.bookinginfo.get('region/hotel', '')
            item['Search Rooms'] = self.bookinginfo.get('number of rooms', '')
            item['Search Adults'] = self.bookinginfo.get('number of adults', '')
            item['Search Children'] = self.bookinginfo.get('number of  kids', '')
            item['Search Date Checkin'] = self.bookinginfo.get('check in date', '')
            item['Search Date Checkout'] = self.bookinginfo.get('check out date', '')

            item['Booking Hotel Name'] = hotel.css('[data-testid="title"]::text').get('')
            item['Booking Location'] = hotel.css('[data-testid="address"] ::text').get('')
            item['Booking Rooms'] = ', '.join(hotel.css('h4[role="link"] ::text').getall())

            item['Booking Adults'] = ''.join(string).split('adults')[0].split(',')[1].strip()
            item['Booking Children'] = ''.join(string).split('children')[0].split('adults,')[1].strip()
            # item['Booking Children'] = ''.join(
            #     [re.search(r'\b(\d+)\s*child\b', x).group(1) if re.search(r'\b(\d+)\s*child\b', x) else '' for x in
            #      string if 'child' in x])

            item['Booking Date Checkin'] = response.css('[data-testid="date-display-field-start"] span::text').get('')
            item['Booking Date Checkout'] = response.css('[data-testid="date-display-field-end"] span::text').get('')
            # item['Booking Nights'] = ''.join(
            #     [re.search(r'\b(\d+)\s*nights\b', x).group(1) if re.search(r'\b(\d+)\s*nights\b', x) else '' for x in
            #      string if 'nights' in x])
            item['Booking Nights'] = ''.join(string).split('night')[0].strip()
            # item['Booking Nights'] = ''.join(
            #     [re.search(r'\b(\d+)\s*nights\b', x).group(1) for x in string if 'nights' in x])
            item['Booking Price'] = hotel.css('[data-testid="price-and-discounted-price"]::text').get('').replace('US$',
                                                                                                                  '')
            item['Booking URL'] = url

            if url in self.scraped_urls:
                continue

            self.booking_scraped_items.append(item)
            self.scraped_urls.append(url)

    def get_key_values_from_file(self, file_path):
        with open(file_path, mode='r', encoding='utf-8') as input_file:
            data = {}

            for row in input_file.readlines():
                if not row.strip():
                    continue

                try:
                    key, value = row.strip().split('==')
                    data.setdefault(key.strip(), value.strip())
                except ValueError:
                    pass

            return data

    def booking_get_cityid(self, response):
        region = self.bookinginfo.get('region/hotel', '')

        data = {
            "query": "karachi",
            "pageview_id": "7ddc3c97c9eb0306",
            "aid": 304142,
            "language": "en-gb",
            "size": 5
        }

        # Extract pageview_id and b_aid values from the response
        script_tag = response.css('script:contains(booking_extra) ::text').get('').split('=')[-1].replace(';',
                                                                                                          '').strip()
        data_dict = json.loads(
            script_tag.replace("'", '"').replace(":", '":').replace(",", ',\n"').replace('"\n', '"').replace('page',
                                                                                                             '"page'))
        pageview_id = data_dict.get('pageview_id', '')
        b_aid = data_dict.get('b_aid', '')

        # Update the data dictionary with extracted values
        data['pageview_id'] = pageview_id
        data['aid'] = b_aid
        data['query'] = region

        # Convert the dictionary to a JSON string
        updated_data = json.dumps(data)

        res = requests.post('https://accommodations.booking.com/autocomplete.json', headers=self.booking_cityid_headers,
                            data=updated_data)
        data = res.json()

        # dest_id = ''.join([x.get('dest_id', '') for x in data.get('results', [{}])])
        dest_id = data.get('results', [{}])[0].get('dest_id', 0)
        dest_type = data.get('results', [{}])[0].get('dest_type', 0)
        ss = data.get('results', [{}])[0].get('label', '')
        content = {
            'ss': ss,
            'dest_id': dest_id,
            'dest_type': dest_type,

        }
        return content

    def booking_query_url(self, response, city_id):
        params = {
            'ss': 'lahore',
            'label': 'gen173nr-1BCAEoggI46AdIM1gEaLUBiAEBmAEJuAEYyAEM2AEB6AEBiAIBqAIEuALP78eqBsACAdICJDNkM2ZlZjkxLTZkNTctNGJkYy1hMDM5LTM4ZjBhMWUxNDk4M9gCBeACAQ',
            'sid': '5f9a5eb7f61da87d92888e838ae0ada0',
            'aid': '304142',
            'lang': 'en-gb',
            'sb': '1',
            'src_elem': 'sb',
            'src': 'index',
            'dest_id': '-2767043',  # get_cityid
            'dest_type': 'city',  # get_cityid
            'ac_position': '0',
            'ac_click_type': 'b',
            'ac_langcode': 'en',
            'ac_suggestion_list_length': '3',
            'search_selected': 'true',
            'search_pageview_id': '710548670e10023c',
            'ac_meta': 'GhA3MTA1NDg2NzBlMTAwMjNjIAAoATICZW46BmxhaG9yZUAASgBQAA==',
            'checkin': '2023-11-22',  # self.bookinginfo
            'checkout': '2023-11-25',  # self.bookinginfo
            'group_adults': '3',  # self.bookinginfo
            'no_rooms': '2',  # self.bookinginfo
            'group_children': '1',  # self.bookinginfo
            'age': '3',  # self.bookinginfo
            # 'age': [
            #     '3',
            #     '3',
            #     '3',
            # ],
            'sb_travel_purpose': 'leisure',
        }
        bookinginfo_ss = self.bookinginfo.get('region/hotel', '')  # location
        label = response.css('script:contains(b_original_url) ::text').re_first(r'b_label:.*').split(':')[1].rstrip(
            ',').strip().replace("'", '')
        sid = response.css('script:contains(b_original_url) ::text').re_first(r'b_sid:.*').split(':')[1].rstrip(
            ',').strip().replace("'", '')
        pageview_id = response.css('script:contains(b_original_url) ::text').re_first(r'pageview_id:.*').split(':')[
            1].rstrip(',').strip().replace("'", '')
        aid = response.css('script:contains(b_original_url) ::text').re_first(r'aid:.*').split(':')[1].rstrip(
            ',').strip().replace("'", '')

        params['label'] = label
        params['sid'] = sid
        params['aid'] = aid
        params['search_pageview_id'] = pageview_id
        # params['ss'] = ss
        params['ss'] = city_id.get('ss', bookinginfo_ss)

        dest_id = city_id.get('dest_id', '')
        params['dest_id'] = dest_id
        dest_type = city_id.get('dest_type', '')
        params['dest_type'] = dest_type
        checkin = self.bookinginfo.get('check in date', '')
        params['checkin'] = checkin
        checkout = self.bookinginfo.get('check out date', '')
        params['checkout'] = checkout
        group_adults = self.bookinginfo.get('number of adults', 0)
        params['group_adults'] = group_adults
        no_rooms = self.bookinginfo.get('number of rooms', 0)
        params['no_rooms'] = no_rooms
        group_children = self.bookinginfo.get('number of  kids', 0)
        params['group_children'] = group_children
        age = self.bookinginfo.get('age of kids', 0)
        params['age'] = age

        # url = f'https://www.booking.com/searchresults.en-gb.html?ss=islamabad&label=gen173nr-1BCAEoggI46AdIM1gEaLUBiAEBmAEJuAEYyAEM2AEB6AEBiAIBqAIEuALP78eqBsACAdICJDNkM2ZlZjkxLTZkNTctNGJkYy1hMDM5LTM4ZjBhMWUxNDk4M9gCBeACAQ&sid=5f9a5eb7f61da87d92888e838ae0ada0&aid=304142&lang=en-gb&sb=1&src_elem=sb&src=index&dest_id=356ttt2c-89d291f5c6e3404d86dd3c1e6ac5dcec-en&dest_type=latlong&place_id=356ttt2c-89d291f5c6e3404d86dd3c1e6ac5dcec-en&latitude=31.0703&longitude=74.9369&ac_position=0&ac_click_type=b&ac_langcode=en&ac_suggestion_list_length=5&search_selected=true&search_pageview_id=710548670e10023c&checkin=2023-11-22&checkout=2023-11-25&group_adults=2&no_rooms=1&group_children=1&age=0&sb_travel_purpose=leisure'
        url = f'https://www.booking.com/searchresults.en-gb.html?ss=islamabad&label={label}&sid={sid}&aid={aid}&lang=en-gb&sb=1&src_elem=sb&src=index&dest_id={dest_id}&dest_type= {dest_type}&ac_position=0&ac_click_type=b&ac_langcode=en&ac_suggestion_list_length=5&search_selected=true&search_pageview_id={pageview_id}&checkin={checkin}&checkout={checkout}&group_adults={group_adults}&no_rooms={no_rooms}&group_children={group_children}&age={age}&sb_travel_purpose=leisure&order=price'
        return url

    def travelor_parse(self, response):
        hotels_selector = self.travelor_selenium_indexpage(response)

        for hotel in hotels_selector:
            url = 'https://www.travelor.com' + hotel.css('.pb-6 a.group::attr(href)').get('')
            price = hotel.css('.text-lg.text-green-500::text').get('').replace('$', '')
            yield Request(url=url, callback=self.travelor_hotel_details, meta={'price': price})

    def travelor_hotel_details(self, response):

        item = OrderedDict()
        script = response.css('script:contains("window.__NUXT__")::text').get('').split(',')
        city_name_string = next((x for x in script if 'city_name' in x), None)
        travelor_location = city_name_string.split(':')[1].replace('"',
                                                                   '') if city_name_string else self.bookinginfo.get(
            'region/hotel', '')

        item['Travelor Hotel Name'] = response.css('h1.inline::text').get('')
        item['Travelor Address'] = response.css('.inline-flex .mx-2::text').get('')
        item['Travelor Location'] = travelor_location
        item['Travelor Rooms'] = self.bookinginfo.get('number of rooms', '')
        item['Travelor Adults'] = self.bookinginfo.get('number of adults', '')
        item['Travelor Children'] = self.bookinginfo.get('number of  kids', '')
        item['Travelor Date Checkin'] = response.url.split('&')[0].split('=')[1]
        item['Travelor Date Checkout'] = response.url.split('&')[1].split('=')[1]
        item['Travelor Nights'] = self.get_travelor_nightstay(response.url)
        item['Travelor Price'] = response.meta.get('price', '')
        item['Travelor URL'] = response.url

        self.travelor_scraped_items.append(item)

    def travelor_selenium_indexpage(self, response):

        if self.driver is None:
            edge_options = EdgeOptions()
            edge_options.use_chromium = True  # Use Chromium-based Edge
            self.driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
            self.driver.maximize_window()

        self.driver.get('https://www.travelor.com/?fid=68193')
        # self.driver.get('https://www.travelor.com/he?fid=68193')
        sleep(3)
        location = self.bookinginfo.get('region/hotel', '')

        input_search = 'destination'
        search_input_field = self.driver.find_element(By.ID, input_search)
        search_input_field.send_keys(location)
        time.sleep(2)
        search_input_field.send_keys(Keys.ARROW_DOWN)  # Move down to the first option
        search_input_field.send_keys(Keys.RETURN)  # Select the option

        time.sleep(2)

        # clicked Search Button
        search_button = 'button.block.relative.h-12 span'
        button = WebDriverWait(self.driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, search_button)))
        button.click()

        # Wait until 'place' is found in the current URL
        WebDriverWait(self.driver, 20).until(EC.url_contains('place'))

        time.sleep(5)

        current_url = self.driver.current_url
        place_id = current_url.split('place/')[1].split('/')[0]
        rooms = int(self.bookinginfo.get('number of rooms', ''))
        date_checkin = self.bookinginfo.get('check in date', '')
        date_checkout = self.bookinginfo.get('check out date', '')
        adult_value = self.bookinginfo.get('number of adults', '')
        adult = ''.join(['a,'] * int(adult_value))
        children = self.bookinginfo.get('number of  kids', '')
        childs = ','.join(['10'] * int(children))
        latitude = current_url.split('latitude')[1].split('&')[0].replace('=', '')
        longitude = current_url.split('longitude')[1].split('&')[0].replace('=', '')
        session_id = current_url.split('session')[1].replace('=', '')

        guests_value = f'{adult} {childs}'

        if rooms > 1:
            guests_value += (rooms - 1) * '%7C'

        url = f'https://www.travelor.com/hotels/place/{place_id}/results?fid=68193&check_in={date_checkin}&check_out={date_checkout}&guests={guests_value}&country=PK&currency=USD&radius=50000&latitude={latitude}&longitude={longitude}&session={session_id}'
        self.driver.get(url)

        # Scroll down to the bottom of the page
        body = self.driver.find_element(By.TAG_NAME, 'body')
        body.send_keys(Keys.END)  # Scroll to the bottom
        time.sleep(3)
        # You can use a loop to scroll several times
        for _ in range(5):  # Scroll 5 times (adjust the number as needed)
            body.send_keys(Keys.END)
            time.sleep(3)

        time.sleep(5)

        html = Selector(text=self.driver.page_source)
        html_selector = html.css('.mt-6 .w-full.pb-6')

        self.driver.quit()

        return html_selector

    def get_travelor_nightstay(self, url):
        url = url
        check_in_date = url.split('?')[1].split('&')[0].split('=')[1]
        check_out_date = url.split('?')[1].split('&')[1].split('=')[1]
        check_in_datetime = datetime.strptime(check_in_date, '%Y-%m-%d')
        check_out_datetime = datetime.strptime(check_out_date, '%Y-%m-%d')

        number_of_nights = (check_out_datetime - check_in_datetime).days

        return number_of_nights

    def close(self, reason):
        self.writetocsv()

    def writetocsv(self):
        headers = ['Search Location', 'Search Rooms', 'Search Adults', 'Search Children',
                   'Search Date Checkin', 'Search Date Checkout',
                   'Booking Hotel Name', 'Booking Location', 'Booking Rooms', 'Booking Adults',
                   'Booking Children', 'Booking Date Checkin', 'Booking Date Checkout',
                   'Booking Nights', 'Booking Price', 'Booking URL',
                   'Travelor Hotel Name', 'Travelor Address', 'Travelor Location', 'Travelor Rooms',
                   'Travelor Adults', 'Travelor Children', 'Travelor Date Checkin', 'Travelor Date Checkout',
                   'Travelor Nights', 'Travelor Price', 'Travelor URL'
                   ]
        wb = Workbook()
        ws = wb.active

        # Write headers to the first row
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        combined_items = []  # Initialize an empty list to store all combined items

        for combined_item in zip_longest(self.booking_scraped_items, self.travelor_scraped_items, fillvalue={}):
            # Combine items into a single row
            combined_item = {**combined_item[0], **combined_item[1]}
            combined_items.append(combined_item)

        # Iterate over rows in combined_items
        for row_num, combined_item in enumerate(combined_items, 2):
            # Iterate over headers
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                # Access the value from the combined_item dictionary and write to the worksheet
                ws[f"{col_letter}{row_num}"] = combined_item.get(header, '_')

        # Save the workbook
        output_file = 'output/Hotels Booking Details.xlsx'
        wb.save(output_file)
        self.logger.info(f'CSV file written: {output_file}')
