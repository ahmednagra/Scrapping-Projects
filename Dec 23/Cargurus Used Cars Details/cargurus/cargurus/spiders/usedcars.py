import glob
import json
import os
from collections import OrderedDict
from datetime import datetime
from math import ceil

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from scrapy import Spider, Request


class UsedcarsSpider(Spider):
    name = "cars"
    start_urls = ['https://www.cargurus.com/']

    headers = ['Make', 'Model', 'Year', 'Title', 'Location', 'Condition', 'Price',
               'Average Market Price', 'Main Image', 'Options',
               'Mileage', 'Trim', 'Color', 'Body Type', 'Fuel Type', 'Engine',
               'Deal Rating', 'Days on Market', 'Drive Train', 'City Fuel Economy',
               'Highway Fuel Economy', 'Combined Fuel Economy', 'Dealer Name',
               'Dealer City', 'Dealer State', 'Dealer Zip', 'Dealer Rating',
               # 'Dealer Reviews Count',
               'Dealer Phone', 'Vin No', 'URL']

    custom_settings = {
        'CONCURRENT_REQUESTS': 3,
        'FEED_EXPORTERS': {'xlsx': 'scrapy_xlsx.XlsxItemExporter'},
        'FEEDS': {
            f'output/Cargurus Used Cars Details.xlsx': {
                'format': 'xlsx',
                'fields': headers,
            }
        },
    }

    def __init__(self, **kwargs):

        super().__init__(**kwargs)
        self.files_rename()
        self.current_scraped_items = []

        self.previously_scraped_items = {item.get('Vin No'): item for item in
                                         self.read_excel_file('output/Previous Cargurus Used Cars Details.xlsx')}

    def parse(self, response, **kwargs):
        companies_value = response.css('#carPickerUsed_makerSelect optgroup option::attr(value)').getall()
        for company in companies_value[12:13]:
            print('Comapnay Value : ', company)
            url = f'https://www.cargurus.com/Cars/l-Used-{company}'
            yield Request(url=url, callback=self.pagination, meta={'company': company})

    def pagination(self, response):
        url = response.url
        script = response.css('script:contains(totalListings)').re_first(r'({.*})')
        data = json.loads(script)
        total_results = data.get('totalListings', 0)
        print('total products are :', total_results)
        if total_results > 1:
            total_pages = ceil(total_results / 48)
            print('Total Pages are :', total_pages)
            for page_number in range(1, total_pages + 1):
                if page_number > 2:
                    break
                offset = page_number * 48
                company = response.meta.get('company', '')
                page_url = f'https://www.cargurus.com/Cars/searchResults.action?entitySelectingHelper.selectedEntity={company}&sourceContext=untrackedExternal_false_0&inventorySearchWidgetType=AUTO&sortDir=ASC&sortType=DEAL_RATING_RPL&shopByTypes=MIX&srpVariation=DEFAULT_SEARCH&nonShippableBaseline=285&offset={offset}&maxResults=48&filtersModified=true'
                yield Request(url=page_url, callback=self.parse_usedcars, dont_filter=True, meta={'url': url})

    def parse_usedcars(self, response):
        try:
            data = json.loads(response.body)

            if data is None or not isinstance(data, (list, dict)):
                self.logger.error("Invalid data format: Unable to parse used cars.")
                return

            for car in data:
                item = OrderedDict()
                make = car.get('makeName', '')
                year = car.get('carYear', '')
                item['Make'] = car.get('makeName', '')
                item['Model'] = car.get('modelName', '').replace(str(make), '').replace(str(year), '').strip()
                item['Year'] = car.get('carYear', '')
                item['Title'] = car.get('listingTitle', '')
                item['Location'] = car.get('sellerCity', '')
                item['Condition'] = 'Used'

                item['Price'] = car.get('priceString', '')
                item['Average Market Price'] = f"${car.get('expectedPrice', '')}"
                item['Main Image'] = car.get('originalPictureData', {}).get('url', '')
                item['Options'] = ', '.join([x for x in car.get('options', '')])

                item['Mileage'] = f"{car.get('mileageString', '')} Miles"
                item['Trim'] = car.get('trimName', '')

                item['Color'] = car.get('exteriorColorName', '') or car.get('normalizedExteriorColor', '')
                item['Body Type'] = car.get('bodyTypeName', '')
                item['Fuel Type'] = car.get('localizedFuelType', '')
                item['Engine'] = car.get('localizedEngineDisplayName', '')

                item['Deal Rating'] = car.get('dealRating', '')
                item['Days on Market'] = car.get('daysOnMarket', '')

                item['Drive Train'] = car.get('driveTrain', '')
                item['City Fuel Economy'] = ' '.join([str(value) for value in list(car.get('cityFuelEconomy', {}).values())])
                item['Highway Fuel Economy'] = ' '.join([str(value) for value in list(car.get('highwayFuelEconomy', {}).values())])
                item['Combined Fuel Economy'] = ' '.join([str(value) for value in list(car.get('combinedFuelEconomy', {}).values())])

                item['Dealer Name'] = car.get('serviceProviderName', '')
                item['Dealer City'] = car.get('sellerCity', '').split(',')[0].strip()
                item['Dealer State'] = car.get('sellerRegion', '')
                item['Dealer Zip'] = car.get('sellerPostalCode', '')
                item['Dealer Rating'] = round(car.get('sellerRating', 0), 2)
                # item['Dealer Reviews Count'] = car.get('reviewCount', 0)
                item['Dealer Phone'] = car.get('phoneNumberString', '')

                item['Vin No'] = car.get('vin', '')
                item['URL'] = response.meta.get('url') + f"#listing={str(car.get('id', ''))}/NONE/DEFAULT"

                yield item

        except Exception as e:
            self.logger.error(f"Error while parsing used cars: {e}")

    # def parse_usedcars(self, response):
    #     try:
    #         data = json.loads(response.body)
    #
    #         if data is None or not isinstance(data, (list, dict)):
    #             self.logger.error("Invalid data format: Unable to parse used cars.")
    #             return
    #
    #         for car in data:
    #             vin_no = car.get('vin', '')
    #             previous_item = self.previously_scraped_items.get(vin_no)
    #
    #             if previous_item:
    #                 self.current_scraped_items.append(previous_item)
    #
    #                 yield previous_item
    #                 continue
    #
    #             id = car.get('id', 0)
    #             Location = car.get('sellerCity', '')
    #             color = car.get('exteriorColorName', '') or car.get('normalizedExteriorColor', '')
    #             body_type = car.get('bodyTypeName', '')
    #
    #             dealer_request = f'https://www.cargurus.com/Cars/detailListingJson.action?inventoryListing={id}&searchZip=&searchDistance=100&inclusionType=DEFAULT'
    #             yield Request(url=dealer_request, callback=self.parse_cardetail,
    #                           meta={'Location': Location, 'color': color, 'body_type': body_type})
    #
    #     except Exception as e:
    #         self.logger.error(f"Error while parsing used cars: {e}")

    def parse_cardetail(self, response):
        data = json.loads(response.body)
        seller_dict = data.get('seller', {})
        car_dict = data.get('listing', {})
        item = OrderedDict()

        item['Make'] = car_dict.get('makeName', '')
        item['Model'] = car_dict.get('listingTitle', '')
        item['Year'] = car_dict.get('year', '')

        item['Location'] = response.meta.get('Location', '')
        item['Condition'] = car_dict.get('vehicleCondition', '')
        item['Price'] = car_dict.get('priceString', '')
        item['Average Market Price'] = car_dict.get('expectedPrice', '')
        item['Main Image'] = car_dict.get('spin', {}).get('spinUrl', '')
        item['Other Images'] = ', '.join([x.get('url', '') for x in car_dict.get('pictures', [])])

        item['Mileage'] = car_dict.get('unitMileage', {}).get('value', 0.0)
        item['Trim'] = car_dict.get('trimName', '')
        item['Color'] = response.meta.get('color', '')
        item['Body Type'] = response.meta.get('body_type', '')
        item['Fuel Type'] = car_dict.get('localizedFuelType', '')
        item['Engine'] = car_dict.get('localizedEngineDisplayName', '')

        item['Deal Rating'] = car_dict.get('dealRatingKey', '')
        item['Accident History Info'] = car_dict.get('vehicleHistory', {}).get('accidentCount', 0)
        item['Days on Market'] = car_dict.get('listingHistory', {}).get('daysOnCarGurus', '')

        item['Dealer Name'] = seller_dict.get('name', '')
        item['Dealer Address'] = ', '.join(seller_dict.get('address', {}).get('addressLines', []))
        item['Dealer Phone'] = seller_dict.get('phoneNumberString', '')
        item['Vin No'] = car_dict.get('vin', '')

        self.current_scraped_items.append(item)
        yield item

    def close(spider, reason):
        spider.comparison_data()

    def get_work_sheet(self, sheet_name, wb, headers):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(headers)

        return sheet

    def comparison_data(self):
        if not self.current_scraped_items:
            return

        current_time = datetime.now()
        formatted_time = current_time.strftime("%Y-%m-%d_%H-%M-%S-%f")[:-2]

        # Specify the file name and output folder
        output_folder = 'output/Reports'
        file_name = f'Sold_Cars_{formatted_time}.xlsx'
        file_path = os.path.join(output_folder, file_name)

        # Create the output folder if it doesn't exist
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Create or load the workbook
        if os.path.isfile(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            wb.remove(sheet)

        # # Create or get sheets
        sold_cars = self.get_work_sheet(sheet_name='previous_records', wb=wb, headers=self.headers)

        # Get the list of VINs from current_scraped_items
        current_scraped_vin = [item.get('Vin No') for item in self.current_scraped_items]

        # Iterate through previous items, and append to sold_cars if VIN not in current_scraped_vin
        for previous_vin, previous_item in self.previously_scraped_items.items():
            if previous_vin not in current_scraped_vin:
                row_values = list(previous_item.values())
                sold_cars.append(row_values)

        # Save the workbook
        wb.save(file_path)

    def files_rename(self):
        all_files = glob.glob('output/*')
        # Check if any file has "previous" in the filename and delete it
        for file_path in all_files:
            if "previous" in file_path.lower():
                os.remove(file_path)

        # Find the remaining file
        remaining_files = glob.glob('output/*')
        for file in remaining_files:
            if 'Previous' not in file and 'Cargurus Used Cars Details' in file:

                # Rename the remaining file by adding "Previous" to the filename
                new_file_path = os.path.join('output', f'Previous {os.path.basename(file)}')
                os.rename(file, new_file_path)
                return

            else:
                return

    def read_excel_file(self, file_path):
        data = []
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

        except Exception as e:
            print(f"Error reading Excel file: {e}")

        return data
