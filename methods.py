import csv
import os
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook


def read_text_file(self):
    file_path = 'input/city_names.txt'

    try:
        with open(file_path, mode='r') as txt_file:
            return [line.strip() for line in txt_file.readlines() if line.strip()]

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        return []
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return []


def write_to_excel(data, sheet_name):
    # Create the directory if it doesn't exist
    output_dir = 'output/'
    os.makedirs(output_dir, exist_ok=True)

    # Save the workbook
    file_name = f'{output_dir}Booking Names_Price {datetime.now().strftime("%d%m%Y")}.xlsx'

    # Create a new workbook or load existing workbook if file already exists
    if os.path.isfile(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()

    # Select the sheet with the spider name or create a new one if it doesn't exist
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)

    # Write the headers if the sheet is empty
    if sheet.max_row == 0:
        headers = ['Name', 'Actual_Price', 'Discounted_Price', 'Date_start', 'Date_end', 'Guests_adult',
                   'Guest_children', 'City']
        sheet.append(headers)

    # Write the data rows
    for row in data:
        flattened_row = [item if not isinstance(item, list) else item[0] for item in row]
        sheet.append(flattened_row)

    try:
        wb.save(file_name)
        print(f"Data saved to {file_name}")
    except Exception as e:
        print(f"An error occurred while saving the data: {str(e)}")


def process_csv_file(self):
    data = []

    try:
        with open(self.input_file, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            data = list(csv_reader)

    except FileNotFoundError:
        print(f"File '{self.input_file}' not found.")
        return
    except Exception as e:
        print(f"An error occurred while reading the file: {str(e)}")
        return

def form_data(self, page_num, category_id):
    return {
        "params": {
            "f_brandno": [],
            "page": page_num,
            "sort": "pageview",
            "f_price": {},
            "category": str(category_id),
            "f_size": [],
            "size": 400,
            "keyword": ""
        },
        "session_id": "71088121b102433db4a2bcdeab689d05",
        "m_no": None
    }

form_params = self.form_data(page_num=1, category_id=category_id)
body=(json.dumps(form_params).encode('utf-8')),
method = 'POST',
